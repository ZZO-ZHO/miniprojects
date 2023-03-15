# 대량 메일 전송
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


wbook = load_workbook('./studyPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active

for i in range(1,wsheet.max_row+1):
    recv_email = wsheet.cell(i, 1).value
    print(recv_email)
    try:
        # ㅅ
        send_mail = 'xstar202@naver.com'
        send_pass = '**********'    # 비밀번호
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        msg = MIMEMultipart()
        msg['Subject'] = '엑셀에서 보내는 메일'
        msg['From'] = send_mail
        msg['To'] = recv_email
        msg.attach(MIMEText('보내는 내용입니다. ㅎㅅㅎ'))

        mail = smtplib.SMTP(smtp_name,smtp_port)
        mail.starttls()
        mail.login(send_mail, send_pass)
        mail.sendmail(send_mail, recv_email, msg.as_string())
        mail.quit()
        print('전송완료')
    except Exception as e:
        print(f'수신메일 - {recv_email}')
        print(f'전송에러 : {e}')